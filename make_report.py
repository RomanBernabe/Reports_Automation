#! python3
#make_report.py - It only works with Excel installed on the machine. 

"""
This script automates the operations report for the daily 8:00am meeting.  It
needs 5 Excel files, downloaded from Odoo:

1.- A file with all the current Open tickets, renamed 'raw_data.xlsx'.
2.- A file with all the Open tickets from the last 24-72 hours, renamed 'last24.xlsx'.
3.- A file with all the Open AND Breached tickets, renamed 'breached.xlsx'.
4.- A csv file with the backup summary. This one is not renamed.
5.- A template report, called 'report.xlsx' which will then be renamed with the
    day of the script execution.

These files need to be on the same folder as the script. The script does the
following:

- Renames the first column on 'raw.xlsx' as 'Ticket#' and cleans said column to 
give only the ticket numbers. 
- Selects only the incidents and copies them to the 'OpenInc' sheet on the report
template.
- Selects only the service requests and copies them to the 'OpenSR' sheet on the 
report template.
- Removes empty rows on the previously created sheets.
- Copies this edited sheet and pastes it to the 'Open' sheet on the report template.
- Copies the data from 'Breached' file and pastes it to the 'BreachedOdoo' sheet
 on the report template.
- Copies the data from 'last24' file and pastes it to the 'Last24c' sheet on the 
report template.
- Copies the only sheet in the csv file to the last sheet on the target report.

It uses python, openpyxl and pywin32. All of them can be downloaded and installed
without admin rights, you only need python for that. 'pip install --user' """


##### Start ############

import openpyxl                      #manipulates Excel
import os                            #core utilities
import shutil                        #let us copy
import time
from pathlib import Path             #let us handle file paths
from win32com.client import Dispatch #manipulates Excel
from datetime import datetime        #gives us the date


# Find all our Excel files to use and put them on variables. Uses the Path library.
pwd = Path.cwd()
source_doc    = pwd / 'raw_data.xlsx'
last24_doc    = pwd / 'last24.xlsx'
breached_doc  = pwd / 'breached.xlsx'
target_report = pwd / 'report_template.xlsx'

get_csv = list(pwd.glob('*.csv'))
backup_jobs =  get_csv[0]

# Define the function that will rename the first column and put the ticket numbers.
def addTicketNumb(file_name):

    # Load the Raw_Data workbook and use the first sheet
    wb = openpyxl.load_workbook(file_name)
    sheet = wb.active

    # Rename first column
    sheet['A1'] = 'Ticket#'

    # Declare variables to use in the for loop
    cell_to_clean = ''
    cleaned_cell = ''

     #Remember, a for loop goes up to the selected range but it doesn't include it,
     #that's why we add one to the max number of rows. Start from 2 to ignore the
     # first row. 
    for i in range(2, sheet.max_row + 1):
        #Assign the corresponding cell value to a variable
        cell_to_clean = sheet['A' + str(i)].value
        #Only get the ticket number of the whole string by list slicing
        cleaned_cell = cell_to_clean[27:32]
        #Put the previous obtained string on the corresponding cell
        sheet['A' + str(i)] = cleaned_cell

    #Save the sheet and return this new file
    wb.save(file_name)

#test addTicketNumb(source_doc)

# This function selects only the incidents from the previous workbook and
# puts them on new sheet in the same workbook. This sheet then will be pasted on
# the final report. This function leaves empty rows. 
def filterIncidents(file_to_filter):
    
    #Open workbook and activate the first sheet.
    wb = openpyxl.load_workbook(file_to_filter) 
    sheet = wb.active 
    #Create a new sheet, named Incidents.  
    receiving_sheet = wb.create_sheet(title='Incidents')

    # Loop adds header to the new sheet
    for columnNum in range(1, sheet.max_column + 1):
        c = sheet.cell(row = 1, column = columnNum)
        receiving_sheet.cell(row = 1, column = columnNum).value = c.value

    
    #Put on a list what types of incidents we are going to select.
    incident_types =['Incident Critical (P1)',
                    'Incident High (P2)',
                    'Incident Normal (P3)',
                    'Incident Low (P4)',
                     False]

    #Loop through all the sheet rows
    for i in range(2, sheet.max_row + 1):
        #On every row, loop through all columns
        for j in range(1, sheet.max_column + 1):
            #Assign the cell value to a variable
            c = sheet.cell(row = i, column = j)
            #If the cell value on column B is one of our designated incident types,
            #  add the value to the newly created 'Incidents' sheet
            if sheet['B' + str(i)].value in incident_types:
                receiving_sheet.cell(row = i, column = j).value = c.value

    #Save the sheet. 
    wb.save(file_to_filter)

#test filterIncidents(source_doc)

# Does the same as the filterIncidents function, but this time it selects only
# the service requests and puts them on a new sheet. 
def filterServiceRequests(file_to_filter):

    wb = openpyxl.load_workbook(file_to_filter) 
    sheet = wb.active
    # Creates a new sheet, named OpenSR. SR = Service Requests.
    receiving_sheet = wb.create_sheet(title='OpenSR')
    
    # Loop add header to new sheet
    for columnNum in range(1, sheet.max_column + 1):
        c = sheet.cell(row = 1, column = columnNum)
        receiving_sheet.cell(row = 1, column = columnNum).value = c.value


    request_types =['Service Request Low',
                    'Service Request Normal']
                    

    for i in range(2, sheet.max_row + 1):
        for j in range(1, sheet.max_column + 1):
            c = sheet.cell(row = i, column = j)
            if sheet['B' + str(i)].value in request_types:
                receiving_sheet.cell(row = i, column = j).value = c.value

    wb.save(file_to_filter)

#test filterServiceRequests(source_doc)


# Previous functions leaves the new sheets with a lot of empy rows. This one
# removes all of them.
def removeEmptyRowsInc(file_to_clean):

    wb = openpyxl.load_workbook(file_to_clean) 
    # Select on what sheet to perform the actions
    sheet = wb['Incidents'] 

    # Loop has to parse from bottom to top. Row numbers don't get altered that way.
    for i in reversed(range(1, sheet.max_row + 1)):
        # We only need to know if the cell in first column is empty to know if 
        # the whole row is empty; if so, remove that row.
        if sheet.cell(i, 1).value is None:
            sheet.delete_rows(i, 1)
    
    wb.save(file_to_clean)

# test removeEmptyRowsInc(source_doc)

# Same as removeEmptyRowsInc, but cleans the SR sheet.
def removeEmptyRowsSR(file_to_clean):

    wb = openpyxl.load_workbook(file_to_clean) 
    sheet = wb['OpenSR'] 

    for i in reversed(range(1, sheet.max_row + 1)):
        if sheet.cell(i, 1).value is None:
            sheet.delete_rows(i, 1)
    
    wb.save(file_to_clean)

#test removeEmptyRowsSR(source_doc)


# Copies the 'Sheet1' sheet from the source_open woorkbook and pastes it on the
# 'Open' sheet on the report template. Uses the Dispatch utility from pywin32 library
def copyOpen(source_wb, target_wb):

    # Instatiates Excel to be able to manipulate it 
    xl = Dispatch("Excel.Application")
    xl.Visible = False  # Remove this line if you don't want Excel to be visible
    xl.DisplayAlerts = False
    xl.Interactive = False
    xl.Application.EnableEvents = False


    # Select what workboos to open and assign them to variables
    wb1 = xl.Workbooks.Open(source_wb)
    wb2 = xl.Workbooks.Open(target_wb)

    # Activate our desired sheet from the first book
    ws1 = wb1.Worksheets('Sheet1')
    ws1.Activate()
    # Select all cells from that sheet
    ws1.UsedRange.Activate()
    # Paste the selection on our desired range
    xl.Selection.Copy(Destination=wb2.Worksheets('Open').Range('A1'))

    # Save the changes and close the app
    wb2.Close(SaveChanges=True)
    #xl.Quit()

#test copyOpen(source_doc, target_report)

# Same as copyOpen, but copies the 'Incidents' sheets from source_doc and pastes
# it on 'OpenInc' on the report template 
def copyInc(source_wb, target_wb):

   
    xl = Dispatch("Excel.Application")
    xl.Visible = False  # Remove this line if you don't want Excel to be visible
    xl.DisplayAlerts = False
    xl.Interactive = False
    xl.Application.EnableEvents = False

    wb1 = xl.Workbooks.Open(source_wb)
    wb2 = xl.Workbooks.Open(target_wb)

    #wb1 = xl.Workbooks.Open(source_doc)
    #wb2 = xl.Workbooks.Open(target_report)
  
    ws1 = wb1.Worksheets('Incidents')
    ws1.Activate()
    ws1.UsedRange.Select() #Use the select method when there are various sheets in wb

    # Here the pasting range changes to not alter the header row on the target 
    xl.Selection.Copy(Destination=wb2.Worksheets('OpenInc').Range('A1'))

     
    wb2.Close(SaveChanges=True)
    #xl.Quit()

#test copyInc(source_doc, target_report)


# Same as copyOpen, but pastes on the 'OpenSR' sheet.
def copyOpenSR(source_wb, target_wb):

   
    xl = Dispatch("Excel.Application")
    xl.Visible = False  # Remove this line if you don't want Excel to be visible
    xl.DisplayAlerts = False
    xl.Interactive = False
    xl.Application.EnableEvents = False

    wb1 = xl.Workbooks.Open(source_wb)
    wb2 = xl.Workbooks.Open(target_wb)

    ws1 = wb1.Worksheets('OpenSR')
    ws1.Activate()
    ws1.UsedRange.Select()

    # Here the pasting range changes to not alter the header row on the target 
    xl.Selection.Copy(Destination=wb2.Worksheets('OpenSR').Range('A1'))
 
    wb2.Close(SaveChanges=True)
    #xl.Quit()

#copyOpenSR(source_doc)

# Same as copyOpen, but pastes the first sheet on the 'Last24C' sheet on the
# report template
def copyLast24(source_wb, target_wb):

 
    xl = Dispatch("Excel.Application")
    xl.Visible = False  # Remove this line if you don't want Excel to be visible
    xl.DisplayAlerts = False
    xl.Interactive = False
    xl.Application.EnableEvents = False

    wb1 = xl.Workbooks.Open(source_wb)
    wb2 = xl.Workbooks.Open(target_wb)

  
    ws1 = wb1.Worksheets('Sheet1')
    ws1.Activate()
    ws1.UsedRange.Activate()

    xl.Selection.Copy(Destination=wb2.Worksheets('Last24C').Range('A1'))

     
    wb2.Close(SaveChanges=True)
    #xl.Quit()

#test copyLast24(source_doc, target_report) 


def copyBreached(source_wb, target_wb):

  
    xl = Dispatch("Excel.Application")
    xl.Visible = False  # Remove this line if you don't want Excel to be visible
    xl.DisplayAlerts = False
    xl.Interactive = False
    xl.Application.EnableEvents = False

    wb1 = xl.Workbooks.Open(source_wb)
    wb2 = xl.Workbooks.Open(target_wb)

  
    ws1 = wb1.Worksheets('Sheet1')
    ws1.Activate()
    ws1.UsedRange.Activate()

    xl.Selection.Copy(Destination=wb2.Worksheets('BreachedOdoo').Range('A1'))

     
    wb2.Close(SaveChanges=True)
    #xl.Quit()

#test copyBreached(source_doc, target_report)


def copyJobs(source_wb, target_wb):

  
    xl = Dispatch("Excel.Application")
    xl.Visible = False  # Remove this line if you don't want Excel to be visible
    xl.DisplayAlerts = False
    xl.Interactive = False
    xl.Application.EnableEvents = False

    wb1 = xl.Workbooks.Open(source_wb)
    wb2 = xl.Workbooks.Open(target_wb)
       
    wb1.Worksheets(1).Copy(Before=wb2.Worksheets(wb2.Worksheets.Count))
    wb2.Worksheets(wb2.Worksheets.Count).Delete()
     
    wb1.Close(SaveChanges=True)
    wb2.Close(SaveChanges=True)
    xl.Quit()

#test copyJobs(backup_jobs, target_report)





def renameReport():
    
    report_headname = 'Daily Operations Report - '
    # Get today's date on the specific format
    current_timestamp = datetime.today().strftime('%m_%d_%Y')
    # Concatenate the head title with the timestamp to have the final title
    report_final_name = report_headname + current_timestamp + '.xlsx'
    # Use shutil's copy method to generate the new report
    shutil.copy(target_report, pwd / report_final_name)

def deleteFiles():

    os.unlink(source_doc)
    os.unlink(last24_doc)
    os.unlink(breached_doc)
    os.unlink(backup_jobs)

def main():

    # Clean first columns phase
    print('Cleaning first columns...')
    addTicketNumb(source_doc)
    addTicketNumb(last24_doc)
    addTicketNumb(breached_doc)

    #Filter and remove empty rows phase
    print('Filtering requests and incidents...')
    filterIncidents(source_doc)
    filterServiceRequests(source_doc)
    removeEmptyRowsInc(source_doc)
    removeEmptyRowsSR(source_doc)


    print('Copying values on to the new report...')
    # Copy-pasting to target workbook phase
    copyOpen(source_doc, target_report)
    copyInc(source_doc, target_report) #doesnt copy
    copyOpenSR(source_doc, target_report) #doesnt copy
    copyLast24(source_doc, target_report)
    copyBreached(source_doc, target_report)
    copyJobs(backup_jobs, target_report)

    print('Adding date to report...')
    renameReport()
    deleteFiles()

    print('Report done. Go ahead and ace that meeting! Until the next time. Bye.')

if __name__ == "__main__":
        main()



